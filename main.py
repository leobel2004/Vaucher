from fpdf import FPDF
import openpyxl


class PDF(FPDF):
    def __init__(self, a, b, c):
        super().__init__()
        self.add_font("Montserrat", style="", fname="font/ofont.ru_Montserrat Alternates.ttf")
        self.add_font("Montserrat", style="I", fname="font/ofont.ru_Montserrat Alternates (1).ttf")
        self.add_font("Montserrat", style="B", fname="font/ofont.ru_Montserrat Alternates (2).ttf")

    def header(self):
        # Rendering logo:
        self.image("IMG_0133.PNG", 50, 0, 160)
        self.ln(40)

    def footer(self):
        # Setting position at 1.5 cm from bottom:
        self.set_y(-20)
        pdf.set_font('Montserrat', 'I', 14)
        pdf.multi_cell(180, 5,
                       "По всем возникающим вопросам просьба обращаться \n по тел. +7(915)146-81-91 Динара, +7(903)267-77-88 Анастасия", new_x="LMARGIN", new_y="NEXT", align ="C")
        # Setting font: helvetica italic 8
        self.set_font("Montserrat", "I", 8)
        # Setting text color to gray:
        self.set_text_color(128)
        # Printing page number
        self.cell(0, 10, f"Страница {self.page_no()}", align="C")


month_list = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
              'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']


book = openpyxl.load_workbook('input.xlsx')
sheet = book.worksheets[0]
for row in sheet.values:
    if row[1] and row[1] != "ФИО":
        Name = row[1]
        Date_of_transfer = row[2]
        Time_of_transfer = row[3]
        Adress_1 = row[4]
        Adress_2 = row[5]
        With = row[6]
        Hotel = row[7]
        Typ = row[8]
        Eat = row[9]
        Enter_Date = row[10]
        Enter_Time = row[11]
        Departure_Date = row[12]
        Departure_Time = row[13]

        pdf = PDF("P", 'mm', 'A4')
        pdf.add_page()
        pdf.set_font("Montserrat", "B", 16)
        pdf.cell(45)
        pdf.multi_cell(100, 10, "ВАУЧЕР на транспортное обслуживание и проживание.", new_x="LMARGIN", new_y="NEXT", align='C')
        pdf.set_font('Montserrat', 'I', 16)
        pdf.cell(10, 10, new_x="LMARGIN", new_y="NEXT")
        pdf.cell(50)
        pdf.cell(100, 10, Name, new_x="LMARGIN", new_y="NEXT", align='C')
        pdf.ln(10)
        pdf.set_font('Montserrat', '', 15)

        for irow in sheet.values:
            if irow[1] == Name and irow[4]:
                text = (str(irow[2])[8:10] + ' ' + month_list[int(str(irow[2])[5:7]) - 1] + ' в ' + str(irow[3])[:5] +
                        ' ' + irow[4] + ' - ' + irow[5])
                if irow[6]:
                    text += " (Совместно с " + irow[6] + ')'
                pdf.multi_cell(180, 10, text, new_x="LMARGIN", new_y="NEXT")
                pdf.ln(5)

        for irow in sheet.values:
            if irow[1] == Name and irow[7]:
                pdf.write_html(f'''<b>Отель: <u>{irow[7]}</u></b>''')
                pdf.cell(10, 5, new_x="LMARGIN", new_y="NEXT")
                text = f"Номер: {irow[8]}"
                if irow[9]:
                    text += f'\nПитание: \n {irow[9]}'
                pdf.multi_cell(180, 10, text, new_x="LMARGIN", new_y="NEXT")
                pdf.ln(4)
                text = ("Даты: " + str(irow[10])[8:10] + ' ' + month_list[int(str(irow[10])[5:7]) - 1] + ' - ' +
                        str(irow[12])[8:10] + ' ' + month_list[int(str(irow[12])[5:7]) - 1] + '\n' + "Заезд с " +
                        str(irow[11])[:5] + '\n' + 'Выезд до '+ str(irow[13])[:5])
                pdf.multi_cell(100, 10, text,new_x="LMARGIN", new_y="NEXT")
                pdf.ln(5)


        pdf.output(f"{Name}.pdf")
